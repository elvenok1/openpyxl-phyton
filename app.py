import io
import requests
from flask import Flask, request, send_file, jsonify
from openpyxl import Workbook

# --- Importaciones clave para la nueva funcionalidad ---
# Importamos la clase Image original con un alias para poder usarla dentro de nuestra clase personalizada
from openpyxl.drawing.image import Image as OriginalImage

# --- CLASE PERSONALIZADA "INTELIGENTE" PARA MANEJAR IMÁGENES ---
# Esta clase reemplazará a la original en el contexto de ejecución del script
class SmartImage(OriginalImage):
    """
    Una clase wrapper que hereda de la clase Image original de openpyxl.
    Intercepta la creación de la imagen para manejar URLs automáticamente.
    """
    def __init__(self, img, **kwargs):
        # Verificamos si el argumento 'img' es una cadena de texto que parece una URL
        if isinstance(img, str) and img.startswith(('http://', 'https://')):
            try:
                # Si es una URL, la descargamos usando requests
                response = requests.get(img)
                response.raise_for_status()  # Lanza un error si la descarga falla (ej. 404)
                
                # Creamos un stream de bytes en memoria con el contenido de la imagen
                image_stream = io.BytesIO(response.content)
                
                # Llamamos al constructor de la clase padre (OriginalImage) 
                # con el stream de la imagen, no con la URL.
                super().__init__(image_stream, **kwargs)
            except requests.exceptions.RequestException as e:
                # Si hay un error de red, lanzamos una excepción clara
                raise IOError(f"No se pudo descargar la imagen desde {img}: {e}")
        else:
            # Si 'img' no es una URL (ej. una ruta de archivo local o un stream ya existente),
            # simplemente dejamos que la clase original se encargue de ello.
            super().__init__(img, **kwargs)


# --- Inicialización de la Aplicación Flask ---
app = Flask(__name__)

# --- Endpoint Principal para Ejecutar Código Openpyxl ---
@app.route('/create-excel-from-script', methods=['POST'])
def create_excel_from_script():
    try:
        python_code = request.get_data(as_text=True)
        if not python_code:
            return jsonify({"error": "No se proporcionó ningún script de Python."}), 400

        wb = Workbook()
        ws = wb.active
        ws.title = "Hoja1 Excel Genius"

        # --- CONTEXTO DE EJECUCIÓN MEJORADO ---
        execution_context = {
            'wb': wb,
            'ws': ws,
            # Módulos y clases de uso común
            'Font': __import__('openpyxl.styles', fromlist=['Font']).Font,
            'PatternFill': __import__('openpyxl.styles', fromlist=['PatternFill']).PatternFill,
            'Alignment': __import__('openpyxl.styles', fromlist=['Alignment']).Alignment,
            'Border': __import__('openpyxl.styles', fromlist=['Border']).Border,
            'Side': __import__('openpyxl.styles', fromlist=['Side']).Side,
            'LineChart': __import__('openpyxl.chart', fromlist=['LineChart']).LineChart,
            'Reference': __import__('openpyxl.chart', fromlist=['Reference']).Reference,
            
            # ¡LA MAGIA OCURRE AQUÍ!
            # Cuando el script del usuario llame a 'Image', en realidad estará llamando a nuestra 'SmartImage'.
            'Image': SmartImage,
        }

        exec(python_code, execution_context)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name='reporte_dinamico.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f"Error durante la ejecución del script: {e}")
        return jsonify({"error": f"Error interno del servidor al ejecutar el script: {str(e)}"}), 500

if __name__ == '__main__':
    app.run(debug=True, port=5001)
